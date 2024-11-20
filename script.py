import csv
from docx import Document
from docx.shared import RGBColor
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
import pandas as pd


def main():
    # --------------------------------1---apple.csv--------------------------------------
    # filename = 'apple.csv'
    #
    # data = []
    # with open(filename, mode='r', newline='', encoding='utf-8') as file:
    #     reader = csv.reader(file)
    #     for row in reader:
    #         data.append(row)
    #
    # for i in range(1, len(data)):
    #     data[i][1] = float(data[i][1]) + 1
    #
    # with open(filename, mode='w', newline='', encoding='utf-8') as file:
    #     writer = csv.writer(file)
    #     writer.writerows(data)
    #
    # print(f'Data has been updated in {filename}')

    # --------------------------------2---example.docx-----------------------------------
    # filename = 'example.docx'
    #
    # doc = Document(filename)
    #
    # for paragraph in doc.paragraphs:
    #     for run in paragraph.runs:
    #         run.font.color.rgb = RGBColor(255, 0, 0)
    #
    # doc.save(filename)

    # --------------------------------3---example.xlsx-----------------------------------

    # filename = 'example.xlsx'
    #
    # workbook = load_workbook(filename)
    # sheet = workbook.active
    #
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         if cell.value:
    #             print(cell.value)
    #             cell.value +=1
    #             cell.font = Font(color="FF0000")
    #
    # workbook.save(filename)
    #
    # print(f'Text color has been changed to red in {filename}')

    # --------------------------------4---input.txt-----------------------------------
    #
    # filename = 'input.txt'
    #
    # with open(filename, mode='r', encoding='utf-8') as file:
    #     content = file.read()
    #     # lines = file.readlines()
    #
    # new_content = content + " Добавлено слово."
    #
    # with open(filename, mode='w', encoding='utf-8') as file:
    #     file.write(new_content)
    #
    # print(f'The word has been added to {filename}')

    # --------------------------------5---sw_templates.json-----------------------------------

    # filename = 'sw_templates.json'
    #
    # with open(filename, mode='r', encoding='utf-8') as file:
    #     data = json.load(file)
    #
    # if isinstance(data['access'], list):
    #     print(data['access'])
    # else:
    #     print("Данные в файле не являются списком.")

    # --------------------------------6---Shuffled.parquet-----------------------------------


    # filename = 'Shuffled.parquet'
    #
    # df = pd.read_parquet(filename)
    #
    # column_name = 'brand'
    # # print(df.columns)
    # if column_name in df.columns:
    #     df[column_name] = df[column_name].astype(str) + " Добавлено слово."
    # else:
    #     print(f"Столбец '{column_name}' не найден в данных.")
    # df.to_parquet(filename, index=False)
    #
    # print(f'The word has been added to each element in the column "{column_name}" in {filename}')


    pass

main()